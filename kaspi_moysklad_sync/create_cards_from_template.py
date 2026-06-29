# -*- coding: utf-8 -*-
"""
Создание карточек товаров на Kaspi из заполненного Excel-шаблона
(kaspi_cards_template.xlsx, который готовит generate_kaspi_template.py).

Берёт строки с «Публиковать = да», собирает по ним запрос products/import и
отправляет на аккаунт A (KASPI_CARDS_API_TOKEN — заведение карточек). Для мебели
каждая строка = отдельная карточка (объединения модификаций нет, см. STATUS).

Фото подставляются из kaspi_images_map.json (готовит export_images_to_kaspi.py),
ключ — SKU; если карты нет, карточка уходит без фото.

Запуск:
  python create_cards_from_template.py --dry-run      # показать, что отправится
  python create_cards_from_template.py --limit 2      # отправить только 2 (тест)
  python create_cards_from_template.py                # отправить все «да»
"""
import argparse
import json
import os

import requests
from openpyxl import load_workbook

import config

HERE = os.path.dirname(os.path.abspath(__file__))
REF_PATH = os.path.join(HERE, "..", "kaspi_category_reference.json")
TEMPLATE_PATH = os.path.join(HERE, "..", "kaspi_cards_template.xlsx")

SOURCE_HEADERS = {
    "Публиковать", "SKU", "Название", "Размер (МС)", "Цвет-исходный (МС)",
    "Высота (МС)", "Ширина (МС)", "Глубина (МС)", "Обивка (МС)", "Категория Kaspi",
    "Бренд",
}
TRUE_WORDS = {"да", "yes", "true", "1", "истина", "+"}


def load_reference():
    with open(REF_PATH, encoding="utf-8-sig") as f:
        return json.load(f)


def load_images_map():
    path = os.path.join(config.KASPI_IMAGES_DIR, "kaspi_images_map.json")
    if not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    # поддерживаем оба формата: {sku: url} и {sku: [url, ...]}
    norm = {}
    for k, v in data.items():
        norm[k] = v if isinstance(v, list) else [v]
    return norm


def _clean_code(header):
    return header[2:].strip() if header.startswith("* ") else header.strip()


def _attr_value(code, raw, ref_cat):
    """Преобразует значение ячейки в значения для Kaspi (список {code,value})."""
    meta = ref_cat.get(code, {})
    if raw is None or str(raw).strip() == "":
        return []
    atype = meta.get("type")
    if atype == "boolean":
        val = str(raw).strip().lower() in TRUE_WORDS
        return [{"code": code, "value": val}]
    if meta.get("multiValued"):
        parts = [p.strip() for p in str(raw).split(",") if p.strip()]
        return [{"code": code, "value": p} for p in parts]
    return [{"code": code, "value": str(raw).strip()}]


def build_items(reference, images_map):
    wb = load_workbook(TEMPLATE_PATH, read_only=True)
    items = []
    issues = []
    for ws in wb.worksheets:
        if ws.title in ("Инструкция", "Списки"):
            continue
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows))
        idx = {h: i for i, h in enumerate(headers) if h}
        for row in rows:
            def cell(name):
                i = idx.get(name)
                return row[i] if i is not None and i < len(row) else None

            if str(cell("Публиковать") or "").strip().lower() not in TRUE_WORDS:
                continue
            sku = cell("SKU")
            category = cell("Категория Kaspi")
            title = cell("Название")
            if not sku or not category:
                issues.append(f"строка без SKU/категории: {sku} / {category}")
                continue
            ref_cat = reference.get(category, {})

            attributes = []
            missing_mandatory = []
            for header, i in idx.items():
                if header in SOURCE_HEADERS:
                    continue
                code = _clean_code(header)
                if code not in ref_cat:
                    continue
                vals = _attr_value(code, row[i] if i < len(row) else None, ref_cat)
                attributes.append((code, vals))
                if ref_cat[code].get("mandatory") and not vals:
                    missing_mandatory.append(code)

            if missing_mandatory:
                issues.append(f"{sku}: не заполнены обязательные: {', '.join(missing_mandatory)}")
                continue

            flat_attrs = [a for _code, vals in attributes for a in vals]
            images = [{"url": u} for u in images_map.get(str(sku), [])]
            brand = cell("Бренд") or config.DEFAULT_BRAND
            items.append({
                "sku": str(sku),
                "title": str(title or sku),
                "brand": str(brand),
                "category": str(category),
                "attributes": flat_attrs,
                "images": images,
            })
    return items, issues


def main():
    parser = argparse.ArgumentParser(description="Создание карточек Kaspi из Excel-шаблона")
    parser.add_argument("--dry-run", action="store_true", help="не отправлять, только показать")
    parser.add_argument("--limit", type=int, default=0, help="отправить только N первых карточек")
    args = parser.parse_args()

    reference = load_reference()
    images_map = load_images_map()
    items, issues = build_items(reference, images_map)

    print(f"Готово к заведению карточек: {len(items)}")
    if issues:
        print(f"Пропущено/проблемы: {len(issues)}")
        for s in issues[:20]:
            print(f"  - {s}")
    if args.limit:
        items = items[:args.limit]
        print(f"Ограничение --limit: отправляю {len(items)}")

    if not items:
        return

    if args.dry_run:
        print("\n[--dry-run] Пример payload первой карточки:")
        print(json.dumps(items[0], ensure_ascii=False, indent=2))
        return

    token = config.KASPI_CARDS_API_TOKEN or config.KASPI_API_TOKEN
    url = f"{config.KASPI_API_URL.replace('/v2', '')}/products/import"
    headers = {"X-Auth-Token": token, "Accept": "application/json", "Content-Type": "text/plain"}
    print(f"\nОтправляю {len(items)} карточек на аккаунт A ({url})...")
    resp = requests.post(url, headers=headers, json=items, timeout=60)
    print(f"  Статус: {resp.status_code}")
    print(f"  Ответ: {resp.text[:1000]}")


if __name__ == "__main__":
    main()
