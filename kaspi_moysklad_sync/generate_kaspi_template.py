# -*- coding: utf-8 -*-
"""
Генератор Excel-шаблона для заведения карточек товаров на Kaspi.

Идея: большую часть данных берём из МойСклад автоматически (название, габариты,
цвет, обивка), а недостающие обязательные характеристики Kaspi (тип, материал,
жёсткость, страна и т.д.) вы дозаполняете в Excel из выпадающих списков —
значения подставлены из реального справочника Kaspi.

Категории (мебель, у них нет *Manufacturer code → каждая модификация = отдельная
карточка):
  - «Стулья (Chairs)»        -> Master - Chairs
  - «Столы (Dinner tables)»  -> Master - Dinner tables
Обеденные группы пока пропускаем (товаров-групп в МойСклад нет).

На вход берёт справочник Kaspi: ../kaspi_category_reference.json
(сформирован запросами к API Kaspi: категории -> характеристики -> значения).

Запуск:  python generate_kaspi_template.py
Результат: ../kaspi_cards_template.xlsx
"""
import argparse
import json
import os
import re

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config
import moysklad_client

# --- пути ---
HERE = os.path.dirname(os.path.abspath(__file__))
REF_PATH = os.path.join(HERE, "..", "kaspi_category_reference.json")
OUT_PATH = os.path.join(HERE, "..", "kaspi_cards_template.xlsx")

# Какие категории Kaspi выводим листами (код Kaspi -> имя листа).
SHEETS = [
    ("Master - Chairs", "Стулья (Chairs)", "chairs"),
    ("Master - Dinner tables", "Столы (Dinner tables)", "tables"),
]

DEFAULT_COUNTRY = "Казахстан"

# Исходные колонки из МойСклад (показываем слева для удобства заполнения).
SOURCE_COLS = ["Публиковать", "SKU", "Название", "Размер (МС)", "Цвет-исходный (МС)",
               "Высота (МС)", "Ширина (МС)", "Глубина (МС)", "Обивка (МС)"]

# Стили
HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
MAND_FILL = PatternFill("solid", fgColor="C6E0B4")   # обязательные Kaspi-поля
SRC_FILL = PatternFill("solid", fgColor="D9D9D9")
PREFILL_FONT = Font(color="1F4E78")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_reference():
    with open(REF_PATH, encoding="utf-8-sig") as f:
        return json.load(f)


def attr_columns(cat_obj):
    """Список (code, meta) атрибутов категории: обязательные сначала."""
    items = list(cat_obj.items())
    items.sort(key=lambda kv: (not kv[1].get("mandatory"), kv[0]))
    return items


def classify(name):
    n = (name or "").lower()
    if re.search(r"стул|табурет|кресл", n):
        return "chairs"
    if re.search(r"стол", n) and not re.search(r"столешк", n):
        return "tables"
    return None


def _norm(s):
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def match_color(values, color_enum):
    """Сопоставляет значения цвета из МойСклад со справочником Furniture*Color.
    Возвращает (matched_csv, raw_csv). matched — то, что нашлось в enum Kaspi."""
    enum_norm = {_norm(v): v for v in color_enum}
    matched, raw = [], []
    for v in values:
        v = (v or "").strip()
        if not v:
            continue
        raw.append(v)
        key = _norm(v)
        if key in enum_norm and enum_norm[key] not in matched:
            matched.append(enum_norm[key])
        else:
            # частичное совпадение (например "светлый орех" содержит "орех")
            for ek, ev in enum_norm.items():
                if ek in key and ev not in matched:
                    matched.append(ev)
                    break
    return ", ".join(matched), ", ".join(raw)


def match_enum(value, enum_values):
    """Единичное сопоставление строки со справочником (для Обивки)."""
    if not value:
        return ""
    key = _norm(value)
    norm = {_norm(v): v for v in enum_values}
    if key in norm:
        return norm[key]
    for ek, ev in norm.items():
        if ek in key or key in ek:
            return ev
    return ""


def product_attr(product, name):
    for a in product.get("attributes", []):
        if a.get("name") == name:
            return a.get("value")
    return None


def has_export_flag(product):
    """Стоит ли у товара галочка «Выгружать в каспи» (флаг наследуется
    модификациями, т.к. своих доп. полей у модификаций нет)."""
    return bool(product_attr(product, config.EXPORT_FLAG_ATTRIBUTE_NAME))


def fetch_moysklad():
    """Возвращает (products_by_id, variants) с нужными полями."""
    products = moysklad_client.list_products()
    by_id = {p["id"]: p for p in products}
    variants = []
    offset, limit = 0, 1000
    while True:
        data = moysklad_client._get("/entity/variant", params={"limit": limit, "offset": offset})
        rows = data.get("rows", [])
        variants.extend(rows)
        if len(rows) < limit:
            break
        offset += limit
    return by_id, variants


def build_rows(products_by_id, variants, require_flag=True):
    """Готовит строки по категориям. Каждая строка — dict исходных полей МойСклад.
    По умолчанию берём только товары с галочкой «Выгружать в каспи»."""
    rows = {"chairs": [], "tables": []}

    def src_from(product, variant=None):
        name = (variant or product).get("name") or product.get("name")
        # SKU: «Артикул» (если заполнен, работает как override) -> иначе «Код».
        # Совпадает с приоритетом в build_assortment_index (фид остатков).
        sku = (variant or product).get("article") or (variant or product).get("code")
        # характеристики цвета берём из модификации, плюс доп.поле Цвет товара
        color_vals = []
        if variant:
            for c in variant.get("characteristics", []):
                if "цвет" in (c.get("name") or "").lower():
                    color_vals.append(c.get("value"))
        pcvet = product_attr(product, "Цвет")
        if pcvet:
            color_vals.append(pcvet)
        size_val = ""
        if variant:
            for c in variant.get("characteristics", []):
                if _norm(c.get("name")) == "размер":
                    size_val = c.get("value")
        if not size_val:
            size_val = product_attr(product, "Размер") or ""
        return {
            "name": name,
            "parent_name": product.get("name") or "",
            "sku": sku,
            "size": size_val,
            "color_vals": color_vals,
            "height": product_attr(product, "Высота") or "",
            "width": product_attr(product, "Ширина") or "",
            "depth": product_attr(product, "Глубина") or "",
            "uphol": product_attr(product, "Обивка") or "",
            "model": product_attr(product, "Модель") or "",
            "brand": product_attr(product, "Бренд") or "",
        }

    # модификации
    for v in variants:
        parent_id = moysklad_client._extract_id(v.get("product", {}).get("meta", {}).get("href", ""))
        product = products_by_id.get(parent_id)
        if not product:
            continue
        if require_flag and not has_export_flag(product):
            continue
        if v.get("article"):
            continue  # «Артикул» заполнен = привязка к существующей карточке Kaspi,
                      # новую карточку не заводим (идёт только в остатки)
        cat = classify(product.get("name"))
        if cat in rows:
            rows[cat].append(src_from(product, v))

    # товары без модификаций (готовые изделия)
    for p in products_by_id.values():
        if (p.get("variantsCount") or 0) > 0:
            continue
        if require_flag and not has_export_flag(p):
            continue
        if product_attr(p, config.KASPI_SKU_ATTRIBUTE_NAME):
            continue  # «Каспи артикулы» заполнено = привязка к существующей карточке
        cat = classify(p.get("name"))
        if cat in rows:
            rows[cat].append(src_from(p))

    return rows


def model_name(src):
    """Имя модели для Kaspi (Chairs*Model и т.п.) — БЕЗ цвета и ОДИНАКОВОЕ для
    всех цветов одной модели, чтобы Kaspi объединил их в одну карточку с выбором
    цвета (объединение мебели идёт по Бренд+Модель, см. пример Arto MODERN)."""
    if src.get("model"):
        return src["model"]
    pn = src.get("parent_name") or src.get("name") or ""
    pn = re.sub(r"\([^)]*\)", "", pn).strip()          # убрать цвет в скобках
    pn = re.sub(r"^(стул(ья)?|табурет(ы)?|кресл[оа]|стол(ик)?(ы)?)\s+", "",
                pn, flags=re.IGNORECASE).strip()        # убрать тип в начале
    return pn or (src.get("parent_name") or "")


# Маппинг исходных полей МойСклад -> коды атрибутов Kaspi (по категориям).
# Возвращает значение для предзаполнения колонки Kaspi или "".
def prefill(cat_key, attr_code, src, color_enum, uphol_enum):
    if attr_code == "Furniture*Color":
        matched, _ = match_color(src["color_vals"], color_enum)
        return matched
    if attr_code == "Furniture*Country":
        return DEFAULT_COUNTRY
    if attr_code.endswith("*Model"):
        return model_name(src)
    if cat_key == "chairs":
        if attr_code == "Chairs*Height":
            return src["height"]
        if attr_code == "Chairs*Width":
            return src["width"]
        if attr_code == "Chairs*Depth":
            return src["depth"]
        if attr_code == "Chairs*Upholster":
            return match_enum(src["uphol"], uphol_enum)
        if attr_code == "Chairs*Type":
            n = (src["name"] or "").lower()
            if "табурет" in n:
                return "табурет"
            if "барный" in n:
                return "барный стул"
            return "стул"
        if attr_code == "Chairs*Number of chairs":
            return 1
    if cat_key == "tables":
        if attr_code == "Dinner tables*Height":
            return src["height"]
        if attr_code == "Dinner tables*Width":
            return src["width"]
        if attr_code in ("Dinner tables*Length", "Dinner tables*Table lingth"):
            return src["depth"]
    return ""


def write_sheet(wb, lists_ws, lists_state, cat_code, sheet_name, cat_key, cat_obj, rows):
    ws = wb.create_sheet(sheet_name)
    attrs = attr_columns(cat_obj)
    color_enum = cat_obj.get("Furniture*Color", {}).get("values", [])
    uphol_code = "Chairs*Upholster" if cat_key == "chairs" else None
    uphol_enum = cat_obj.get(uphol_code, {}).get("values", []) if uphol_code else []

    n_src = len(SOURCE_COLS)
    headers = list(SOURCE_COLS) + ["Категория Kaspi", "Бренд"] + [
        (("* " if a[1].get("mandatory") else "") + a[0]) for a in attrs
    ]
    ws.append(headers)
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci)
        c.font = HDR_FONT
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if ci <= n_src:
            c.fill = SRC_FILL
            c.font = Font(bold=True, size=10)
        elif ci <= n_src + 2:            # «Категория Kaspi», «Бренд»
            c.fill = HDR_FILL
        else:
            a = attrs[ci - n_src - 3]
            c.fill = MAND_FILL if a[1].get("mandatory") else HDR_FILL
            if not a[1].get("mandatory"):
                c.font = HDR_FONT

    first_kaspi_col = n_src + 3  # +Категория +Бренд
    for src in rows:
        line = [
            "да",
            src["sku"], src["name"], src["size"],
            ", ".join([x for x in src["color_vals"] if x]),
            src["height"], src["width"], src["depth"], src["uphol"],
            cat_code, src["brand"],
        ]
        for a_code, _meta in attrs:
            line.append(prefill(cat_key, a_code, src, color_enum, uphol_enum))
        ws.append(line)

    # выпадающие списки для enum-колонок (через лист «Списки»)
    max_row = ws.max_row
    for idx, (a_code, meta) in enumerate(attrs):
        if meta.get("type") != "enum":
            continue
        values = meta.get("values", [])
        if not values:
            continue
        col_letter = get_column_letter(first_kaspi_col + idx)
        list_col = _put_list(lists_ws, lists_state, values)
        rng = f"'Списки'!${list_col}$2:${list_col}${len(values) + 1}"
        dv = DataValidation(type="list", formula1=rng, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max(max_row, 2)}")

    # ширины колонок + заморозка шапки
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 34
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.column_dimensions["C"].width = 34


def _put_list(lists_ws, state, values):
    """Кладёт список значений в новую колонку листа «Списки», возвращает букву."""
    col = state["next_col"]
    state["next_col"] += 1
    col_letter = get_column_letter(col)
    lists_ws.cell(row=1, column=col, value=f"список {col_letter}")
    for i, v in enumerate(values, start=2):
        lists_ws.cell(row=i, column=col, value=v)
    return col_letter


def write_instructions(wb):
    ws = wb.create_sheet("Инструкция", 0)
    lines = [
        "Шаблон для заведения карточек товаров на Kaspi (из МойСклад).",
        "",
        "Как заполнять:",
        "1. Серые колонки слева (МС) — данные из МойСклад, для справки. Не обязательны.",
        "2. Зелёные колонки с «*» — ОБЯЗАТЕЛЬНЫЕ характеристики Kaspi. Заполните все.",
        "3. Синие колонки — необязательные характеристики Kaspi.",
        "4. У enum-колонок есть выпадающий список допустимых значений Kaspi.",
        "5. Колонка «Публиковать»: оставьте «да» для строк, которые заводим на Kaspi,",
        "   поставьте «нет» для остальных.",
        "6. Часть значений предзаполнена из МойСклад (габариты, обивка, цвет, страна) —",
        "   проверьте их. Цвет: где значение было кодом ткани, поле осталось пустым —",
        "   выберите цвет вручную из списка (Furniture*Color).",
        "7. Габариты в МойСклад местами в разных единицах — приведите к см/мм по Kaspi.",
        "",
        "После заполнения карточки заводятся скриптом create_cards_from_template.py",
        "на аккаунт Kaspi A (заведение карточек).",
    ]
    for i, t in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=t)
    ws.column_dimensions["A"].width = 90


def main():
    parser = argparse.ArgumentParser(description="Генератор Excel-шаблона карточек Kaspi")
    parser.add_argument("--all", action="store_true",
                        help="включить все товары (игнорировать флажок «Выгружать в каспи»)")
    args = parser.parse_args()
    require_flag = not args.all

    print("Читаю справочник категорий Kaspi...")
    ref = load_reference()

    print("Загружаю товары и модификации из МойСклад...")
    products_by_id, variants = fetch_moysklad()
    print(f"  товаров: {len(products_by_id)}, модификаций: {len(variants)}")

    rows = build_rows(products_by_id, variants, require_flag=require_flag)
    print(f"  строк: стулья={len(rows['chairs'])}, столы={len(rows['tables'])}"
          + ("" if not require_flag else "  (только с галочкой «Выгружать в каспи»)"))
    if require_flag and not rows["chairs"] and not rows["tables"]:
        print("  ВНИМАНИЕ: нет товаров с галочкой «Выгружать в каспи».")
        print("  Проставьте галочки нужным товарам в МойСклад, либо запустите с --all для предпросмотра.")

    wb = Workbook()
    wb.remove(wb.active)  # убираем дефолтный лист
    lists_ws = wb.create_sheet("Списки")
    lists_ws.sheet_state = "hidden"
    lists_state = {"next_col": 1}

    for cat_code, sheet_name, cat_key in SHEETS:
        cat_obj = ref.get(cat_code)
        if not cat_obj:
            print(f"  ВНИМАНИЕ: нет {cat_code} в справочнике — лист пропущен")
            continue
        write_sheet(wb, lists_ws, lists_state, cat_code, sheet_name, cat_key,
                    cat_obj, rows[cat_key])

    write_instructions(wb)
    wb.save(OUT_PATH)
    print(f"\nГотово: {OUT_PATH}")


if __name__ == "__main__":
    main()
